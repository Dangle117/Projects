from pathlib import Path
from typing import List
from models.student import Student
import tempfile
import shutil
import csv


def save_students(filename: str, students: List[Student]) -> bool:
    """
    Atomically write the students list to a CSV file (9 columns).

    Format:
      Header: student_id,name,birth_year,major,<DEFAULT_SUBJECTS...>,gpa

    Implementation details:
    - Writes to a temporary file then moves it into place to avoid partial writes.
    - Returns True on success, False on failure (caller can inspect console logs).
    """
    try:
        path = Path(filename)
        # if relative, use current working dir
        if not path.is_absolute():
            path = Path.cwd() / path
        path.parent.mkdir(parents=True, exist_ok=True)

        # write to a temp file then replace atomically
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", delete=False, newline="") as tmp:
            writer = csv.writer(tmp)
            # header row (Excel-friendly)
            header = ["student_id", "name", "birth_year", "major"] + Student.DEFAULT_SUBJECTS + ["gpa"]
            writer.writerow(header)
            for s in students:
                writer.writerow(s.to_row())
            tmp_path = Path(tmp.name)
        shutil.move(str(tmp_path), str(path))
        return True
    except Exception as ex:
        # Simple logging for failure; caller can present a UI message.
        print("File save error:", ex)
        return False


def load_students(filename: str) -> List[Student]:
    """
    Read students from CSV or Excel workbook and return a list of Student objects.

    Notes:
    - Excel reading requires openpyxl; if missing, function returns an empty list and prints a hint.
    - Header detection is heuristic (looks for common header words); both header and non-header files supported.
    - Malformed rows are skipped with a console message.
    """
    students: List[Student] = []
    try:
        path = Path(filename)
        if not path.is_absolute():
            path = Path.cwd() / path

        # --- Excel (.xlsx) support ---
        if path.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            try:
                import openpyxl
            except Exception:
                print("openpyxl is required to load Excel files. Install with: pip install openpyxl")
                return students

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)

            # Peek first row to decide if it's a header row
            first = next(rows, None)
            if first is None:
                return students

            hdr_candidates = [str(c).strip().lower() for c in first if c is not None]
            header_like = any(h in ("student_id", "id", "name", "birth", "birth year", "major", "gpa") for h in hdr_candidates)

            data_rows = [] if header_like else [first]
            data_rows.extend(rows)

            for lineno, row in enumerate(data_rows, start=1):
                # skip empty rows
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                try:
                    # Normalize to length 9
                    cells = list(row) + [""] * max(0, 9 - len(row))
                    # Use Student.from_row to construct
                    s = Student.from_row(cells[:9])
                    students.append(s)
                except Exception as e:
                    print(f"Skipping malformed Excel row {lineno}: {e}")
                    continue
            return students

        # --- CSV/text fallback ---
        if not path.exists():
            return students
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            rows = list(reader)
            if not rows:
                return students
            # Detect header row by checking common header keywords
            first = rows[0]
            hdr = [str(c).strip().lower() for c in first if c is not None]
            header_like = any(h in ("student_id", "id", "name", "birth", "birth year", "major", "gpa") for h in hdr)
            data_rows = rows[1:] if header_like else rows
            for lineno, row in enumerate(data_rows, start=1):
                if not row or all(not str(c).strip() for c in row):
                    continue
                try:
                    # Normalize to length 9 and create Student
                    s = Student.from_row(row[:9])
                    students.append(s)
                except Exception as e:
                    print(f"Skipping malformed CSV row {lineno}: {e}")
                    continue
    except Exception as ex:
        print("File load error:", ex)
    return students


def export_xlsx(filename: str, students: List[Student]) -> None:
    """
    Export students to an .xlsx workbook.

    Requires openpyxl. Raises ImportError if not installed.
    Writes a header row then one row per student using Student.to_row().
    """
    try:
        import openpyxl
        from openpyxl import Workbook
    except Exception:
        raise ImportError("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    # header matching CSV format
    header = ["student_id", "name", "birth_year", "major"] + Student.DEFAULT_SUBJECTS + ["gpa"]
    ws.append(header)
    for s in students:
        ws.append(s.to_row())
    # save workbook (overwrites atomically via temp file not necessary here)
    ws.parent.save(filename)